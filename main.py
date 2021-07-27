import praw
import json
import openpyxl as reader
from pathlib import Path
from Data import StorageData
from Writer import Writer
from Debugger import Debug
import os.path


# checks if the given subreddit exists
def check_sub(subreddit, reddit):
    exists = True
    try:
        reddit.subreddits.search_by_name(subreddit, exact=True)
    except ValueError:
        exists = False
    return exists


# checks if the reddit data file exists, if not makes one
def check_reddit_data_file(file):
    if os.path.isfile(file) == bool(False):
        temp = Writer(file, 'data')
        temp.close()


# uses user.json to connect to file, change the name of userdefault.json and put your values in to use
def open_json_file(file):
    j_file = open(file)
    return json.load(j_file)


def connect_to_reddit(j_data):
    return praw.Reddit(client_id=j_data['id'], client_secret=j_data['secret'],
                       username=j_data['username'], password=j_data['password'],
                       user_agent=j_data['agent'])

# code needs to be moved into a few more functions/classes, but functionally it is all there
def main():
    # sets debug mode (is very simple atm)
    debug = Debug(True)
    print("Enter a subreddit:")
    subreddit = input()

    j_data = open_json_file('user.json')

    # connects to the reddit account using the json file
    reddit = connect_to_reddit(j_data)

    if check_sub(subreddit, reddit):
        sub = reddit.subreddit(subreddit)

        # currently set to a limit of 100, put any value you like or limit=None if you do not want a limit
        top = sub.top(limit=100)

        # sets up the storage structure used to hold all the data
        storage = StorageData()

        file = Path('', 'reddit_data.xlsx')

        check_reddit_data_file(file)

        # opens the excel file and then reads the values from it, as they need to be gathered each time
        book = reader.load_workbook(file)
        sheet = book.active

        # adds the data from the excel file to the storage structure
        for row in sheet.iter_rows(max_row=sheet.max_row):
            i = 0
            identify = {}
            name = {}
            for cell in row:
                if i == 0:
                    identify = cell.value
                elif i == 1:
                    name = cell.value
                elif i == 2:
                    storage.entry(identify, name, cell.value)
                i += 1

        debug.print('data from excel file gathered has been set')

        # sets up the output class that writes to the excel file
        output = Writer(file, "data")

        # adds the gathered data from submission to the storage structure
        for submission in top:
            storage.entry(submission.id, submission.title, submission.score)

        debug.print('data collected from subreddit')

        key_list = list(storage.keys())

        # writes the values from the storage structure to the excel file
        for key in key_list:
            output.write(key, storage.title(key), storage.score(key))

        debug.print('data written to excel file')

        output.close()

        debug.print('program complete')
    else:
        print(subreddit, 'was not found', sep=' ')


if __name__ == "__main__":
    main()
